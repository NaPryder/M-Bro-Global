import os
import traceback
from openpyxl import load_workbook
from datetime import datetime

def readrow(row:list):
    return [cell_value(cell) for cell in row]

def cell_value(cell):
    if cell.value is None:
        return ""
    if type(cell.value) == str:
        return  cell.value.strip()
    return cell.value

def extract_log(txt ,mode='a', encode='utf-8', isstart=False):
    print(txt)
    now = datetime.now().strftime('%Y-%m-%d %H%M%S')
    try:
        with open(PATH.FileLog, mode, encoding=encode) as txtf:
            if not isstart:
                txtf.write(f'''"{now}"","{str(txt)}"\n''')
            else:
                txtf.write(f"Time,Log\n")

    except:
        pass

class PATH:
    ymd = datetime.now().strftime("%Y-%m-%d")
    ROOT = os.path.dirname(os.path.dirname(__file__))   
    #/M_Bro_Global
    MASTER_PATH = os.path.join(ROOT, "Conf", "Master")
    INPUT_PATH = os.path.join(ROOT, "Data", "Input")
    OUTPUT_PATH = os.path.join(ROOT, "Data", "Output")
    TEMP_PATH = os.path.join(ROOT,"Temp")

    FileMasterConfigInput = os.path.join(MASTER_PATH, "ConfigInput.xlsx")
    FileLog = os.path.join(TEMP_PATH, f"Log_{ymd}.txt")

    CHECK_FOLDERS = [INPUT_PATH, OUTPUT_PATH, TEMP_PATH]


    def __init__(self) -> None:
        self.__start_log()
        for folder in PATH.CHECK_FOLDERS:
            try:
                if not os.path.exists(folder):
                    os.makedirs(folder)
            except :
                error = traceback.format_exc()
                extract_log(error)
                return error
    
    def __start_log(self):
        if not os.path.exists(PATH.FileLog):
            extract_log("",mode='w',isstart=True)

class MasterConfig(PATH):

    def __init__(self) -> None:
        self.list_header_UNIVERSITY = []
        self.list_header_PROGRAM_NAME = []
        self.list_header_GENDER = []
        self.list_header_AMT = []
        self.list_header_YEAR = []
        self.list_header_FAC_NAME = []
        self.INDEX_University = 0   #Column B
        self.INDEX_PROGRAM_NAME = 1 #Column C
        self.INDEX_GENDER = 2       #Column D
        self.INDEX_AMT = 3          #Column E
        self.INDEX_YEAR = 4         #Column F
        self.INDEX_FAC_NAME = 5     #Column G

        #====sheet API
        self.api_Key = ""
        self.url = ""
        self.year = ""

    def get_Api_config(self):
        try:
            if os.path.exists(PATH.FileMasterConfigInput):
                wb = load_workbook(PATH.FileMasterConfigInput)
                ws = wb['API']

                self.api_Key =  ws.cell(row=2,column=2).value
                url = ws.cell(row=3,column=2).value
                self.year = ws.cell(row=4,column=2).value

                self.url = f"{url}&api-key={self.api_Key}"

                wb.close()
            else:
                raise(f"Error not foound file {PATH.FileMasterConfigInput}")
        except:
            error = traceback.format_exc()
            extract_log(error)
            return error

    def get_Header_list(self):
        if os.path.exists(PATH.FileMasterConfigInput):
            wb = load_workbook(PATH.FileMasterConfigInput)
            ws = wb['Header']

            self.__get_column_index(ws)
            self.__read(ws)

            wb.close()
        else:
            self.list_header_UNIVERSITY = ['UNIV_NAME_TH']
            self.list_header_PROGRAM_NAME = ['PROGRAM_NAME']
            self.list_header_GENDER = ['GENDER_NAME', 'GENDER']
            self.list_header_AMT = ['amount', 'ผู้สำเร็จการศึกษา', 'จำนวนผู้สำเร็จการศึกษา']
            self.list_header_YEAR = ['AYEAR', 'A_YEAR', 'YEAR']
            self.list_header_FAC_NAME = ['FAC_NAME']

    def __get_column_index(self, ws):
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=2):
            row = readrow(row)
            try:
                self.INDEX_University = row.index("University")
                self.INDEX_PROGRAM_NAME = row.index("PROGRAM_NAME")
                self.INDEX_GENDER = row.index("Gender")
                self.INDEX_AMT = row.index("AMT")
                self.INDEX_YEAR = row.index("Year")
                self.INDEX_FAC_NAME = row.index("FAC_Name")
            except:
                error = traceback.format_exc()
                print(error)
                pass

    def __read(self, ws):
        for row in ws.iter_rows(min_row=2, min_col=2):
            row = readrow(row)
            # print(row)

            self.list_header_UNIVERSITY = self.__add_data_to_list(self.list_header_UNIVERSITY, row[self.INDEX_University])
            self.list_header_PROGRAM_NAME = self.__add_data_to_list(self.list_header_PROGRAM_NAME, row[self.INDEX_PROGRAM_NAME])
            self.list_header_GENDER = self.__add_data_to_list(self.list_header_GENDER, row[self.INDEX_GENDER])
            self.list_header_AMT = self.__add_data_to_list(self.list_header_AMT, row[self.INDEX_AMT])
            self.list_header_YEAR = self.__add_data_to_list(self.list_header_YEAR, row[self.INDEX_YEAR])
            self.list_header_FAC_NAME = self.__add_data_to_list(self.list_header_FAC_NAME, row[self.INDEX_FAC_NAME])

    def __add_data_to_list(self, add_to_list:list, data:str):
        if data and data not in add_to_list:
            add_to_list.append(data)
        return add_to_list

class Header:

    def __init__(self) -> None:
        self.idx_PROGRAM_NAME = 0
        self.idx_AMT = 0
        self.idx_UNIVERSITY = 0
        self.idx_GENDER = 0
        self.idx_YEAR = 0
        self.idx_FAC_NAME = 0
        self.master = MasterConfig()
        self.master.get_Header_list()
        self.UNIVERSITY = self.master.list_header_UNIVERSITY
        self.PROGRAM_NAME = self.master.list_header_PROGRAM_NAME
        self.GENDER = self.master.list_header_GENDER
        self.AMT = self.master.list_header_AMT
        self.YEARS = self.master.list_header_YEAR
        self.FAC_NAME = self.master.list_header_FAC_NAME

    def get_index(self, header):
        self.idx_PROGRAM_NAME = self.__get_index(header, self.PROGRAM_NAME)
        self.idx_AMT = self.__get_index(header, self.AMT)
        self.idx_UNIVERSITY = self.__get_index(header, self.UNIVERSITY)
        self.idx_GENDER = self.__get_index(header, self.GENDER)
        self.idx_YEAR = self.__get_index(header, self.YEARS)
        self.idx_FAC_NAME = self.__get_index(header, self.FAC_NAME)

    def __get_index(self, header, config_header):
        idx = 5
        extract_log(f"header = {header}")
        for header_config in config_header:
            for header_check in header:

                if header_config.strip().upper() == header_check.strip().upper():
                    extract_log(f"header_config = {header_config} header_check:{header_check}")

                    try:
                        idx = header.index(header_config)
                        extract_log(f"idx = {idx}")
                        break
                    except:
                        exit()
        return idx 

class API_Config:

    def __init__(self) -> None:
        self.master = MasterConfig()
        error = self.master.get_Api_config()
        self.api_key = self.master.api_Key
        self.request_url = self.master.url
        self.year = self.master.year


if __name__ == '__main__':


    pass