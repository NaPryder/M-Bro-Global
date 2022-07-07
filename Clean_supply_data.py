import os
import re
import csv
from Config import Header, PATH, extract_log
from openpyxl import Workbook

def get_number(value:str):
    try:
        if value == '' :
            return 0
        else:
            return int(value)
    except:
        pass
    return 0

def get_year_by_name(filename:str):
    regex = re.search(r"^\w+.*(\d{4}).*.csv", filename)
    if regex:
        year = regex.group(1)
        try:
            year = get_number(year) - 543
        except:
            pass
        print(f"year : {year}")
        return year
    else:
        return None

def read_csv(csvfile_read, year, encoding_from='utf-8-sig'):
    all_row = []
    d_cleanup = {}

    header_obj = Header()
    pattern = r"^[A-Za-z0-9\(\)\-\&\#\$\@\!\*\+\/\{\}\[\]\%\^\_\=\,\.\?]+$"

    with open(csvfile_read, encoding=encoding_from) as myFile:  
        reader = csv.reader(myFile)

        n=0
        for row in reader:
            new_row = []
            if n ==0:
                header_obj.get_index(row)
            n+=1

            if len(row) >=5 and n>=1:

                university = row[header_obj.idx_UNIVERSITY]
                program_name = row[header_obj.idx_PROGRAM_NAME]
                gender = row[header_obj.idx_GENDER]
                amt = row[header_obj.idx_AMT]
                faculty = row[header_obj.idx_FAC_NAME]

                if year is None:
                    year = row[header_obj.idx_YEAR]
                    year = get_number(year)+ 1
                    year = year - 543       #change year to Common Era year 
                    year = f"{year}"

                if program_name not in d_cleanup:
                    d_cleanup[program_name] = {'amt':0 , 'faculty':list()}
                d_cleanup[program_name]['amt'] += get_number(amt) 

                if faculty not in d_cleanup[program_name]['faculty']:
                    d_cleanup[program_name]['faculty'].append(faculty)

                if re.search(pattern, program_name.replace(' ','').strip()) : continue

                new_row = [year, university, program_name, gender, amt]
                all_row.append(new_row)    

            if 2<= n <= 7: 
                extract_log(f"old row: {row}")
                extract_log(f"new row: {new_row}")
                
                
    return d_cleanup, all_row

def check_and_save_excel(wb, filesave):
    if os.path.exists(filesave):
        try:
            os.remove(filesave)
        except:
            pass
    wb.save(filesave)
    print(f'save file : {filesave}')

def write_new_data(d_data:dict, file_save):

    wb = Workbook()
    ws = wb.active
    ws.title = 'Program_Name'

    ws.cell(row=1, column=1).value = 'PROGGRAM_NAME'
    ws.cell(row=1, column=2).value = 'AMOUNT'
    ws.cell(row=1, column=3).value = 'FAC_LIST'


    for at_row, (program_name , d_val) in enumerate(d_data.items(), start=2):
        
        ws.cell(row=at_row, column=1).value = program_name
        ws.cell(row=at_row, column=2).value = d_val['amt']
        ws.cell(row=at_row, column=3).value = str(d_val['faculty'])

    wb.save(file_save)
    print(f"save file: {file_save}")
    wb.close()

def write_file_clean_data(all_row:list, file_save, is_Write_header:bool, encode='utf-8-sig'):

    with open(file_save, 'a', newline='\n', encoding=encode) as newfile:  
        writer = csv.writer(newfile, delimiter=',')

        if is_Write_header:
            writer.writerow(['Year', 'University', 'Program', 'Gender', 'Amount'])
            is_Write_header = False

        for row in all_row:
            if row:
                writer.writerow(row)
    return is_Write_header

def Main_write_supply():
    Path = PATH()
    Input_folder = Path.INPUT_PATH
    Output_folder = Path.OUTPUT_PATH
    is_Write_header = True
    saveFileCSV = os.path.join(Output_folder, f"Clean data.csv")
    for file in os.listdir(Input_folder):
        csv_file = os.path.join(Input_folder, file)
        csv_file_cleanup = os.path.join(Output_folder, f"Clean {file.replace('.csv', '.xlsx')}")

        year = get_year_by_name(file)

        extract_log(f"csv_file : {csv_file} year:{year}")

        if os.path.exists(csv_file) and '.csv' in file.lower():
            d_cleanup, all_row = read_csv(csv_file, year)
            
            is_Write_header = write_file_clean_data(all_row=all_row, file_save=saveFileCSV, is_Write_header=is_Write_header)
            write_new_data(d_data=d_cleanup, file_save=csv_file_cleanup)
#Test 
def test_header(header):
    header_obj = Header()
    header_obj.get_index(header)
    print(f" FAC {header_obj.idx_FAC_NAME}")
    print(f" PRO {header_obj.idx_AMT}")
    print(f" AMT {header_obj.idx_PROGRAM_NAME}")


if __name__ == '__main__':
    Main_write_supply()
    pass